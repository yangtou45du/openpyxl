#!/usr/bin/python
# -*- coding: cp936 -*-
# -*- coding: encoding -*-
from openpyxl import load_workbook
import openpyxl
#读取文件
wb=load_workbook("testData1.xlsx")#默认打开文件，可读写

#获取工作表
sheetNames=wb.sheetnames#h获取所有sheet名称
#print(sheetNames)
#a_sheet=wb['Sheet1']#根据sheet名获取sheet
#print a_sheet
#title=a_sheet.title#获取title名
#print(title)
sheet=wb.active#获取当前正在显示的sheet
#print sheet
#获取单元格值
b4=sheet['B4']
print b4.column#获取列
print(b4.row)#获取行
print b4.value#获取单元格值
b4_too=sheet.cell(row=4,column=2)
print(b4_too.value)#获取单元格只
print(sheet.max_row)#获取最大行
print(sheet.max_column)#获取最大列
#sheet.rows每一行的数据,每一行是元祖，sheet.columns每一列，每个元祖是每一列的单元格
for row in sheet.rows:#因为按行，所以返回A1, B1, C1这样的顺序
     for cell in row:
        print cell.value
for column in sheet.columns:# A1, A2, A3这样的顺序
    for cell in column:
        print cell.value
#获取某行的数据
for cell in list(sheet.rows)[2]:
    print cell.value
#获取任意区间的单元格,获得了以A1为左上角，B3为右下角矩形区域的所有单元格
for i in range(1,4):
    for j in range(1,3):
        print(sheet.cell(row=i,column=j).value)
#切片
for row_cell in sheet['A1':'B3']:
   for cell in row_cell:
       print cell.value
for cell in sheet['A1':'B3']:
    print(cell)

#根据字母获得列号，根据列号返回字母
from openpyxl.utils import get_column_letter, column_index_from_string
#根据列的数字返回字母
print get_column_letter((2))#B
print(column_index_from_string('D'))#4

#将数据写入excel
from openpyxl import Workbook
wb=Workbook()#新建工作表，还没保存
sheet=wb.active#激活工作表
sheet.title="test"#直接赋值就可以改工作表的名称
print(wb.sheetnames)#获取表名
wb.create_sheet('data',index=1)# 被安排s到第二个工作表，index=0就是第一个位置
print(wb.sheetnames)#获取表名
#wb.remove(sheet)# 删除某个工作表
#del wb[sheet]# 删除某个工作表
sheet['C1']='good'# 直接给单元格赋值就行
wb.save("testData1.xlsx")
sheet['B9']='=AVERAGE(B2:B8)'# B9处写入平均值,读取的时候需要加上data_only=True这样读到B9返回的就是数字，如果不加这个参数，返回的将是公式本身'=AVERAGE(B2:B8)'
wb.save("testData1.xlsx")
# 添加一行
row = [1 ,2, 3, 4, 5]
sheet.append(row)

# 添加多行
rows = [
    ['Number', 'data1', 'data2'],
    [2, 40, 30],
    [3, 40, 25],
    [4, 50, 30],
    [5, 30, 10],
    [6, 25, 5],
    [7, 50, 10],
]
list(zip(*rows))#按列写入

# out
'''
[('Number', 2, 3, 4, 5, 6, 7),
 ('data1', 40, 40, 50, 30, 25, 50),
 ('data2', 30, 25, 30, 10, 5, 10)]
 '''
